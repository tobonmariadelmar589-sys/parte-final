"""
utils/exportar.py
Permite exportar información a Excel y PDF con filtros aplicados.
Usa openpyxl para Excel y reportlab para PDF.
"""
from datetime import datetime
from tkinter import filedialog, messagebox


def exportar_excel(titulo: str, encabezados: list, filas: list, filtro_info: str = ""):
    """
    Exporta datos a un archivo .xlsx con formato profesional.
    Incluye: encabezado del bufete, fecha, filtros y estilos de color.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar como Excel",
            initialfile=f"{titulo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not ruta:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = titulo[:30]

        # Estilos de color
        fill_navy = PatternFill("solid", fgColor="1E2761")
        fill_gold = PatternFill("solid", fgColor="C9A84C")
        fill_par  = PatternFill("solid", fgColor="F2F4F8")
        borde = Border(
            left  =Side(style='thin'), right =Side(style='thin'),
            top   =Side(style='thin'), bottom=Side(style='thin')
        )
        centro = Alignment(horizontal="center", vertical="center")

        # Fila 1: título del reporte
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=len(encabezados))
        c = ws.cell(row=1, column=1,
                    value=f"JUSTICIA & ASOCIADOS — {titulo.upper()}")
        c.font      = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        c.fill      = fill_navy
        c.alignment = centro

        # Fila 2: fecha y filtros aplicados
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2,   end_column=len(encabezados))
        info = f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        if filtro_info:
            info += f"   |   Filtro: {filtro_info}"
        ws.cell(row=2, column=1, value=info).alignment = centro

        # Fila 3: encabezados de columnas
        for col, enc in enumerate(encabezados, 1):
            c = ws.cell(row=3, column=col, value=enc)
            c.font      = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill      = fill_gold
            c.alignment = centro
            c.border    = borde

        # Filas de datos con color alterno
        for i, fila in enumerate(filas):
            for col, val in enumerate(fila, 1):
                c = ws.cell(row=i + 4, column=col, value=val)
                c.border    = borde
                c.alignment = Alignment(vertical="center")
                if i % 2 == 0:
                    c.fill = fill_par

        # Ajustar ancho de columnas automáticamente
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(ruta)
        messagebox.showinfo("Exportacion exitosa", f"Excel guardado en:\n{ruta}")

    except ImportError:
        messagebox.showerror("Libreria faltante",
                             "Instala openpyxl:\npip install openpyxl")
    except Exception as e:
        messagebox.showerror("Error al exportar Excel", str(e))


def exportar_pdf(titulo: str, encabezados: list, filas: list, filtro_info: str = ""):
    """
    Exporta datos a un archivo .pdf con formato institucional.
    Usa reportlab para generar el documento.
    """
    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        )

        ruta = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="Guardar como PDF",
            initialfile=f"{titulo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
        if not ruta:
            return

        doc = SimpleDocTemplate(
            ruta, pagesize=landscape(A4),
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=2*cm,    bottomMargin=2*cm
        )

        elems = []
        navy  = colors.HexColor("#1E2761")
        gold  = colors.HexColor("#C9A84C")

        # Título y subtítulo
        elems.append(Paragraph("JUSTICIA & ASOCIADOS",
            ParagraphStyle("t", fontSize=16, fontName="Helvetica-Bold",
                           textColor=navy, alignment=1, spaceAfter=4)))
        elems.append(Paragraph(titulo.upper(),
            ParagraphStyle("s", fontSize=12, fontName="Helvetica-Bold",
                           textColor=gold, alignment=1, spaceAfter=4)))

        # Info de exportación y filtros
        info = f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        if filtro_info:
            info += f"   |   Filtro: {filtro_info}"
        elems.append(Paragraph(info,
            ParagraphStyle("i", fontSize=9, fontName="Helvetica",
                           textColor=colors.grey, alignment=1, spaceAfter=10)))
        elems.append(Spacer(1, 0.3*cm))

        # Tabla de datos
        ancho = (landscape(A4)[0] - 3*cm) / len(encabezados)
        data  = [encabezados] + [list(f) for f in filas]
        tabla = Table(data, colWidths=[ancho] * len(encabezados), repeatRows=1)

        estilo = TableStyle([
            ("BACKGROUND",    (0, 0), (-1,  0), navy),
            ("TEXTCOLOR",     (0, 0), (-1,  0), gold),
            ("FONTNAME",      (0, 0), (-1,  0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1,  0), 9),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",      (0, 1), (-1, -1), 8),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ])
        # Filas alternas
        for i in range(1, len(data)):
            if i % 2 == 0:
                estilo.add("BACKGROUND", (0, i), (-1, i),
                           colors.HexColor("#F2F4F8"))
        tabla.setStyle(estilo)
        elems.append(tabla)

        # Pie de página
        elems.append(Spacer(1, 0.5*cm))
        elems.append(Paragraph(f"Total de registros: {len(filas)}",
            ParagraphStyle("p", fontSize=8, fontName="Helvetica",
                           textColor=colors.grey, alignment=2)))

        doc.build(elems)
        messagebox.showinfo("Exportacion exitosa", f"PDF guardado en:\n{ruta}")

    except ImportError:
        messagebox.showerror("Libreria faltante",
                             "Instala reportlab:\npip install reportlab")
    except Exception as e:
        messagebox.showerror("Error al exportar PDF", str(e))
